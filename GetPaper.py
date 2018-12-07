from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait

from openpyxl import load_workbook

#For OSX
driver = webdriver.Chrome('/usr/local/bin/chromedriver')
#driver = webdriver.Chrome('C:/Users/kpark/AppData/Local/Programs/Python/chromedriver.exe')

driver.get('http://www.riss.kr/index.do')

#검색어는 지정주어야 함
#학위논문 대상
def academy_riss():
    handler = driver.window_handles

    for i in range(0,4):
        try:
            driver.switch_to_window(handler[i])
            query_input = driver.find_element_by_xpath("//input[@id='basicQuery']")
            
        except Exception as e:
            print('Closing Popups...')
            driver.close()

    driver.switch_to_window(handler[0])

    # Input of Keywords & Search
    sleep(2)
    query = "uber"
    query_input.send_keys(query)
    search_btn = driver.find_element_by_xpath("//input[@src='/main/images/sc_btn.gif']")
    search_btn.click()
 
    # 학위논문
    graduate_sec = driver.find_element_by_xpath("//a[contains(text(), '학위논문')]")
    graduate_sec.click()

    #100개씩 보여주기
    show_hundred = driver.find_element_by_xpath("//a[contains(text(), '100개씩 출력')]")
    show_hundred.click()

def scrap_academy():
    title_candidates = driver.find_elements_by_xpath("//p[@class='txt']") #전체 돌아가는 숫자
    val = 0 #저자, 소속을 위한 임시 변수

    for element in range(0, len(title_candidates)):
        paper_title = title_candidates[element].text

        info_temp = driver.find_elements_by_xpath("//span[@class='etc']/a")

        print(val)
        paper_author = info_temp[val].text
        val = val + 1
        print(val)
        paper_association = info_temp[val].text
        val = val + 1 
        print(val)

        try:
            year_temp = driver.find_elements_by_xpath("//span[@class='etc']")[element].text
            one_tmp = year_temp.split(' ')
            print(one_tmp)
            two_tmp = one_tmp[2].split(',')
            print(two_tmp)
            paper_year = two_tmp[1][1:-1]
        
        except Exception as e:
            print("Not on its format1")

        try:
            year_temp = driver.find_elements_by_xpath("//span[@class='etc']")[element].text
            one_tmp = year_temp.split(' ')
            print(one_tmp)
            two_tmp = one_tmp[1].split(',')
            print(two_tmp)
            paper_year = two_tmp[1][1:-1]
        except Exception as e:
            print("Not on its format2") 

        try:
            year_temp = driver.find_elements_by_xpath("//span[@class='etc']")[element].text
            one_tmp = year_temp.split(' ')
            print(one_tmp)
            two_tmp = one_tmp[4].split(',')
            print(two_tmp)
            paper_year = two_tmp[1][1:-1]

        except Exception as e:
            print("Not on its format2") 

        # Put infos on Excel
        workbook = load_workbook("/Users/Kyungho/Desktop/PaperScraper/temp.xlsx")
        worksheet = workbook.active
        
        title_idx = 1
        author_idx = 2
        association_idx = 3
        year_idx = 4

        row_idx = element + 1

        worksheet.cell(row=row_idx, column=title_idx, value=paper_title)
        worksheet.cell(row=row_idx, column=author_idx, value=paper_author)
        worksheet.cell(row=row_idx, column=association_idx, value=paper_association)    
        worksheet.cell(row=row_idx, column=year_idx, value=paper_year)
        workbook.save("/Users/Kyungho/Desktop/PaperScraper/temp.xlsx")


#검색어는 지정주어야 함
#국내학술지 대상
def paperwork_riss():
    handler = driver.window_handles

    for i in range(0,2): #팝업창 개수(n) + 1만큼 range(0,n+1)돌려줘야 함
        try:
            driver.switch_to_window(handler[i])
            query_input = driver.find_element_by_xpath("//input[@id='basicQuery']")
            
        except Exception as e:
            print('Closing Popups...')
            driver.close()

    driver.switch_to_window(handler[0])

    # Input of Keywords & Search
    query = "uber"
    query_input.send_keys(query)
    search_btn = driver.find_element_by_xpath("//input[@src='/main/images/sc_btn.gif']")
    search_btn.click()
 
    # 학위논문
    graduate_sec = driver.find_element_by_xpath("//a[contains(text(), '국내학술지논문')]")
    graduate_sec.click()

    #100개씩 보여주기
    show_hundred = driver.find_element_by_xpath("//a[contains(text(), '100개씩 출력')]")
    show_hundred.click()

    sleep(1)


def scrap_paperwork():
    title_candidates = driver.find_elements_by_xpath("//p[@class='txt']") #전체 돌아가는 숫자를 정의하기 위해
    url_checkpoint = driver.current_url #검색결과 화면을 체크
     
    print(len(title_candidates))

    for element in range(0,len(title_candidates)):

        print("New Paper!")
        sleep(2)
        
        try:
            graduate_paper_list = driver.find_elements_by_xpath("//p[@class='txt']/a")
            graduate_paper_list[element].click()
            
            #Title
            title_temp = driver.find_element_by_xpath("//div[@class='vTop02']/p[@class='tit']")
            paper_title = title_temp.text

            #Sub infos - make as grouped body
            info_temp = driver.find_elements_by_xpath("//p[@class='w56']")

            #Author
            paper_author = info_temp[1].text

            #Association
            paper_association = info_temp[4].text

            #Year
            list_temp = driver.find_elements_by_xpath("//ul[@class='report']/li/p")
            paper_year = list_temp[9].text

            #URL
            url_temp = driver.find_element_by_xpath("//div[@class='vTop02']/p[@class='copy']/a")
            paper_url = url_temp.text

            print(paper_title)
            print(paper_author)
            print(paper_association)
            print(paper_year)
            print(paper_url)

            driver.get(url_checkpoint)
            sleep(1)

        except Exception as e:
            print('Minor Error has ocured.')

        # Put infos on Excel
        workbook = load_workbook("/Users/Kyungho/Desktop/PaperScraper/temp.xlsx")
        worksheet = workbook.active
        
        title_idx = 1
        author_idx = 2
        association_idx = 3
        year_idx = 4
        url_idx = 5

        row_idx = element + 1

        worksheet.cell(row=row_idx, column=title_idx, value=paper_title)
        worksheet.cell(row=row_idx, column=author_idx, value=paper_author)
        worksheet.cell(row=row_idx, column=association_idx, value=paper_association)    
        worksheet.cell(row=row_idx, column=year_idx, value=paper_year)
        worksheet.cell(row=row_idx, column=url_idx, value=paper_url)
        workbook.save("/Users/Kyungho/Desktop/PaperScraper/temp.xlsx")


def search_list_by_clicking():
    #전체 페이퍼 항목들을 보여줌
    graduate_paper_list = driver.find_elements_by_xpath("//p[@class='txt']")
    
    for element in range(0, len(graduate_paper_list)):
        print(element)
        graduate_paper_list = driver.find_elements_by_xpath("//p[@class='txt']")
        graduate_paper_list[element].click()
        sleep(1)

        #제목
        paper_title = driver.find_element_by_xpath("//p[@id='vTop02_tit']").text

        #저자
        author_temp = driver.find_elements_by_xpath("//a[@class='text']")
        '''
        # Test Code
        for i in range(0,len(author_temp)):
            print(i)
            print(author_temp[i].text)
        '''
        paper_author = author_temp[0].text
        sleep(1)

        #발행년도
        year_tmp = driver.find_elements_by_xpath("//ul[@class='report']/li/p")
        paper_year = year_tmp[7].text
        sleep(1)

        #URL
        paper_url = driver.find_element_by_xpath("//p[@class='copy']/a").text

        print(paper_title)
        print(paper_author)
        print(paper_year)
        print(paper_url)

def test_excel():
    workbook = load_workbook("/Users/Kyungho/Desktop/PaperScraper/temp.xlsx")
    worksheet = workbook.active
    
    row_idx = element + 1
    title_idx = 1
    author_idx = 2
    association_idx = 3
    year_idx = 4

    worksheet.cell(row=row_idx, column=title_idx, value='제목')
    worksheet.cell(row=row_idx, column=author_idx, value='저자')
    worksheet.cell(row=row_idx, column=association_idx, value='소속')
    worksheet.cell(row=row_idx, column=year_idx, value='연도')
    workbook.save("/Users/Kyungho/Desktop/PaperScraper/temp.xlsx")

    row_idx = row_idx + 1


#academy_riss()
#scrap_academy()

paperwork_riss()
scrap_paperwork()
