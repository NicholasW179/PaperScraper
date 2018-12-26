from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait

from openpyxl import load_workbook

'''
RISS 사이트에서 논문 제목과 저자를 가지고 있을때 구체적인 정보들을 (학술지.. 연도 등)
가져올 수 있는 자동화 스크립트
'''

def string_to_list(val):
    result = val.split()
    return result


def from_name_to_detail():

    #For OSX
    driver = webdriver.Chrome('/usr/local/bin/chromedriver')
    #driver = webdriver.Chrome('C:/Users/kpark/AppData/Local/Programs/Python/chromedriver.exe')

    #Opening Workbook
    workbook = load_workbook("/Users/Kyungho/Desktop/PaperScraper/target.xlsx")
    worksheet = workbook.active

    row_count = worksheet.max_row #마지막 row의 숫자

    for row_num in range(2,row_count):

        try:
            #row는 계속 바꾸고 column은 고정값
            title_temp = worksheet.cell(row=row_num, column=3).value #길이조정 필요 value 넣으면 값을 업데이트 가능
            author_temp = worksheet.cell(row=row_num, column=4).value 

            title = ' '.join(string_to_list(title_temp)[:3])
            author = string_to_list(author_temp)[0]

            print(title)
            print(author)

            #Riss 접근 부분
            driver.get('http://www.riss.kr/index.do')

            handler = driver.window_handles

            for i in range(0,2):
                try:
                    driver.switch_to_window(handler[i])
                    query_input = driver.find_element_by_xpath("//input[@id='basicQuery']")

                except Exception as e:
                    print('Closing Popups...')
                    driver.close()

            driver.switch_to_window(handler[0])

            deep_search = driver.find_element_by_xpath("//a[@title='상세검색']")

            try:
                deep_search.click()
            except Exception as e:
                deep_search.send_keys(Keys.ENTER)

            title_query = driver.find_element_by_xpath("//input[@id='keyword1']")
            title_query.send_keys(title) #title

            author_query = driver.find_element_by_xpath("//input[@id='keyword2']")
            author_query.send_keys(author) #author

            #검색버튼
            search_btn = driver.find_element_by_xpath("//div[@id='sch3']/input[@alt='검색'][@type='image']")

            try:
                search_btn.click()
            except Exception as e:
                search_btn.send_keys(Keys.ENTER)

            identifiers = driver.find_elements_by_xpath("//h2[@class='tit']")
            target = identifiers[0].text

            if '학위논문' in target:
                print('학위논문 입니다.')

                paper_temp = driver.find_elements_by_xpath("//p[@class='txt']/a")
                paper_temp[0].click()

                info_temp = driver.find_elements_by_xpath("//p[@class='w56']")

                #전공분야
                major = info_temp[3].text

                keyword_temp = driver.find_elements_by_xpath("//p/a[@class='text']")
                keyword_box = []
                
                for i in keyword_temp[3:]:
                    keyword_box.append(i.text)

                #키워드(주제)
                keywords = ', '.join(keyword_box)

            elif '국내학술지' in target:
                print('국내학술지 입니다.')
                paper_temp = driver.find_elements_by_xpath("//p[@class='txt']/a")
                paper_temp[0].click()

                info_temp = driver.find_elements_by_xpath("//p[@class='w56']")

                #전공분야
                major = info_temp[2].text

                keyword_temp = driver.find_elements_by_xpath("//p/a[@class='text']")
                keyword_box = []
                
                for i in keyword_temp[6:-1]:
                    keyword_box.append(i.text)

                #키워드(주제)
                keywords = ', '.join(keyword_box)

            else:
                print('해당되지 않습니다')

            worksheet.cell(row=row_num, column=1, value=major)
            worksheet.cell(row=row_num, column=2, value=keywords)

            workbook.save("/Users/Kyungho/Desktop/PaperScraper/target.xlsx")
            print('success!')


        except Exception as e:
            print('End of the Content')

def open_original_paper():
    pass
