from selenium import webdriver
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.action_chains import ActionChains
from time import sleep
from selenium.webdriver.support.ui import WebDriverWait

from openpyxl import load_workbook


#For OSX
driver = webdriver.Chrome('/usr/local/bin/chromedriver')

#driver = webdriver.Chrome('C:/Users/kpark/AppData/Local/Programs/Python/chromedriver.exe')
driver.get('https://dict.naver.com/name-to-roman/translation/?where=name')
sleep(1)

#row_idx = 1

workbook = load_workbook("/Users/Kyungho/Desktop/PaperScraper/name.xlsx")
worksheet = workbook.active
#all_rows = worksheet.rows

for row_idx in range(1,188):
    name_query = driver.find_element_by_xpath("//input[@id='query']")

    kr_name = worksheet.cell(row=row_idx, column=1).value #한국이름

    name_obj = kr_name #이름 체크
    name_query.send_keys(name_obj)

    search_btn = driver.find_element_by_xpath("//input[@src='/static/btn_action3.gif']")
    search_btn.click()

    multiple_results = driver.find_elements_by_xpath("//td[@class='cell_engname']")

    result = multiple_results[0]
    txt_result = result.text
    txt_list = txt_result.split()
    final_result = txt_list[1] + ' ' + txt_list[0]
    print(type(final_result))
    print(final_result)

    eg_name = worksheet.cell(row=row_idx, column=2, value=final_result)

    row_idx = row_idx + 1

    workbook.save("/Users/Kyungho/Desktop/PaperScraper/name.xlsx")


    driver.get('https://dict.naver.com/name-to-roman/translation/?where=name')
    sleep(1)